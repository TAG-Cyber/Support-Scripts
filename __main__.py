import mysql.connector
from selenium import webdriver
from openpyxl import load_workbook
import os
from time import sleep

db_host = ''

print("Connecting to the DB")

connection = mysql.connector.connect(user='', password='', host=db_host, port=, database='vendors')

cursor = connection.cursor()

print("connection cursor found")

query = ("SELECT name, website FROM vendors")

vendor_dict = {}

cursor.execute(query)

for (name, website) in cursor:
  vendor_dict[name.lower()] = website

cursor.close()

print("Cursor closed")

connection.close()

print("Connection closed")

wb = load_workbook(filename='vendors.xlsx')

sheet = wb['vendors']

browser = webdriver.Firefox()

browser.delete_all_cookies()

for row in range(2,sheet.max_row):
  if(sheet.cell(row,1).value is None):
    break
  vendor_actual_name = sheet.cell(row,1).value
  vendor_name = sheet.cell(row,1).value.lower()
  file_name = vendor_name + '.png'
  source = file_name
  dst = 'firefox/' + file_name 
  try:
    if (vendor_dict[vendor_name] != ''):
      browser.get(vendor_dict[vendor_name])
      sleep(1)
      browser.get_screenshot_as_file(file_name)
      os.rename(source, dst)
      url_cell = 'B'+ str(row)
      update_cell = 'C' + str(row)
      print(url_cell)
      sheet[url_cell] = vendor_dict[vendor_name]
      sheet[update_cell] = "Yes"
  except Exception as ex:
    sheet['C' + str(row)] = "No"
    print(ex)
    print(vendor_name + "-> Not Found")

browser.quit()

wb.save(filename='vendors.xlsx')

wb.close()
